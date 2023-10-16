import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd

file_path = r"C:\Users\Mihir\Downloads\Delegate Applications CHIREC MUN 2023   (2).xlsx" 
df = pd.read_excel(file_path)

column_name = 'Tes' 

red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

duplicates = df[df.duplicated(subset=column_name, keep='first')]

workbook = openpyxl.load_workbook(file_path)

sheet = workbook.active

for index, row in duplicates.iterrows():
    row_index = index + 2  
    for col_num, cell_value in enumerate(row):
        col_letter = openpyxl.utils.get_column_letter(col_num + 1)
        cell = sheet[col_letter + str(row_index)]
        cell.fill = red_fill

output_file_path = 'output_file.xlsx' 
workbook.save(output_file_path)
